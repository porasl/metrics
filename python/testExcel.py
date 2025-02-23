import configparser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import pandas as pd

# Read configuration from config.ini
config = configparser.ConfigParser()
config.read('config.ini')

# Fetch the values from the config file
startDate = config['DEFAULT']['startDate']
endDate = config['DEFAULT']['endDate']
storeId = config['DEFAULT']['storeId']
collectionId = config['DEFAULT']['collectionId']

# Print out the values for testing
print(f"Start Date: {startDate}")
print(f"End Date: {endDate}")
print(f"Store ID: {storeId}")
print(f"Collection ID: {collectionId}")

# Generate a new workbook and select the active worksheet
workbook = Workbook()
sheet = workbook.active

data = [
    ["E0000005353", "12/08/2024", "USA434254", "false", "Retail", 32.0],
    ["E0000005354", "12/08/2024", "USA434255", "false", "Business Affiliate", 12.0],
    ["E0000005353", "12/08/2024", "BR434256", "true", "Consumer Retail", 22.0],
    ["E0000005355", "12/08/2024", "BR434257", "false", "Consumer Retail", 132.0],
    ["E0000005355", "12/08/2024", "US434257", "true", "Consumer Retail", 22.0],
    ["E0000005355", "12/08/2024", "UK434257", "false", "Consumer Retail", 32.0],
    ["E0000005356", "12/08/2024", "PU434258", "false", "Business Affiliate", 322.0],
    ["E0000005357", "12/08/2024", "SWE434259", "false", "Retail", 42.0],
    ["E0000005358", "12/08/2024", "FIN434260", "false", "Retail", 412.0]
]

for row in data:
    sheet.append(row)

# Save the initial workbook with data
workbook.save("Sample.xlsx")

# Read the data into a pandas DataFrame
df = pd.DataFrame(data, columns=["Id", "Creation Date", "AccountId", "Guest", "Account Type", "Total"])

# Append "Guest" to the Account Type if the Guest column contains 'true'
df["Account Type"] = df.apply(lambda x: f"{x['Account Type']} Guest" if x['Guest'] == "true" else x['Account Type'], axis=1)

# Perform the aggregation
aggregated_df = df.groupby(["Creation Date", "Account Type"]).agg({"Total": "sum"}).reset_index()

# Check if all dates are the same
unique_dates = aggregated_df["Creation Date"].unique()
if len(unique_dates) == 1:
    date = unique_dates[0]
    # Remove the "Creation Date" column if all dates are the same
    aggregated_df = aggregated_df.drop(columns=["Creation Date"])
else:
    date = None

# Convert Total to currency format
aggregated_df["Total"] = aggregated_df["Total"].apply(lambda x: "${:,.2f}".format(x))

# Add a new sheet for the summary
summary_sheet = workbook.create_sheet(title="Summary")

# Write the date as a header if applicable
row_num = 1
if date:
    summary_sheet.cell(row=row_num, column=1, value=f"Date: {date}")
    row_num += 1

# Write the aggregated data to the Summary sheet
for index, row in aggregated_df.iterrows():
    summary_sheet.cell(row=row_num, column=1, value=row["Account Type"]).alignment = Alignment(horizontal='left')
    summary_sheet.cell(row=row_num, column=2, value=row["Total"])
    row_num += 1

# Calculate the summary of aggregated total values
summary_total = aggregated_df["Total"].str.replace('[$,]', '', regex=True).astype(float).sum()
summary_text = f"Summary of Aggregated Total: ${summary_total:,.2f}"

# Write the summary total
summary_sheet.cell(row=row_num, column=1, value=summary_text)

# Function to adjust the column widths to fit the content
def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Adjust column widths for both sheets
adjust_column_width(sheet)
adjust_column_width(summary_sheet)

# Save the workbook with the adjusted column widths
workbook.save("Sample.xlsx")

print("Aggregated data has been written to the 'Summary' sheet, and column widths have been adjusted.")

