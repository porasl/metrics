from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import pandas as pd

# Generate a new workbook and select the active worksheet
workbook = Workbook()
sheet = workbook.active

data = [
    ["E0000005353", "12/08/2024", "US434254", "false", "Retail", 32.0],
    ["E0000005354", "12/08/2024", "US434255", "false", "Business Affiliate", 12.0],
    ["E0000005353", "12/08/2024", "US434256", "true", "Consumer Retail", 22.0],
    ["E0000005355", "12/08/2024", "US434257", "false", "Consumer Retail", 132.0],
    ["E0000005355", "12/08/2024", "US434257", "true", "Consumer Retail", 22.0],
    ["E0000005355", "12/08/2024", "US434257", "false", "Consumer Retail", 32.0],
    ["E0000005356", "12/08/2024", "US434258", "false", "Business Affiliate", 322.0],
    ["E0000005357", "12/08/2024", "US434259", "false", "Retail", 42.0],
    ["E0000005358", "12/08/2024", "US434260", "false", "Retail", 412.0]
]

for row in data:
    sheet.append(row)

# Save the initial workbook with data
workbook.save("Sample.xlsx")

# Read the data into a pandas DataFrame
df = pd.DataFrame(data, columns=["Id", "Creation Date", "AccountId", "Guest", "Account Type", "Total"])

# Append "Guest" to the Account Type if the Guest column contains 'true'
df["Account Type"] = df.apply(lambda x: f"{x['Account Type']} Guest" if x['Guest'] == "true" else x['Account Type'], axis=1)

# Perform the aggregation
aggregated_df = df.groupby(["Creation Date", "Account Type"]).agg({"Total": "sum"}).reset_index()

# Check if all dates are the same
unique_dates = aggregated_df["Creation Date"].unique()
if len(unique_dates) == 1:
    date = unique_dates[0]
    # Remove the "Creation Date" column if all dates are the same
    aggregated_df = aggregated_df.drop(columns=["Creation Date"])
else:
    date = None

# Convert Total to currency format
aggregated_df["Total"] = aggregated_df["Total"].apply(lambda x: "${:,.2f}".format(x))

# Add a new sheet for the summary
summary_sheet = workbook.create_sheet(title="Summary")

# Write the date as a header if applicable
row_num = 1
if date:
    summary_sheet.cell(row=row_num, column=1, value=f"Date: {date}")
    row_num += 1

# Write the aggregated data to the Summary sheet
for index, row in aggregated_df.iterrows():
    summary_sheet.cell(row=row_num, column=1, value=row["Account Type"]).alignment = Alignment(horizontal='left')
    summary_sheet.cell(row=row_num, column=2, value=row["Total"])
    row_num += 1

# Calculate the summary of aggregated total values
summary_total = aggregated_df["Total"].str.replace('[$,]', '', regex=True).astype(float).sum()
summary_text = f"Summary of Aggregated Total: ${summary_total:,.2f}"

# Write the summary total
summary_sheet.cell(row=row_num, column=1, value=summary_text)

# Function to adjust the column widths to fit the content
def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Adjust column widths for both sheets
adjust_column_width(sheet)
adjust_column_width(summary_sheet)

# Add a bar chart to the summary sheet
# Reference data for the chart
categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=row_num-1)
values = Reference(summary_sheet, min_col=2, min_row=2, max_row=row_num-1)

# Create a bar chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Total by Account Type"
chart.y_axis.title = 'Total Value'
chart.x_axis.title = 'Account Type'

# Set the x-axis labels to be aligned on the right side of the bars
chart.x_axis.label_align = "right"

# Enable data labels to show the values on top of the bars
data_labels = DataLabelList()
data_labels.showVal = True
chart.dataLabels = data_labels

# Add data to the chart
chart.add_data(values, titles_from_data=False)
chart.set_categories(categories)

# Position the chart below the table
summary_sheet.add_chart(chart, f"A{row_num + 2}")

# Save the workbook with the adjusted column widths and added chart
workbook.save("Sample.xlsx")

print("Aggregated data and bar chart with total values on the bars have been written to the 'Summary' sheet, and column widths have been adjusted.")
